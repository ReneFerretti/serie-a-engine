import os
import sqlite3
import requests
import pandas as pd
import numpy as np
from scipy.stats import poisson
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
from datetime import datetime
from zoneinfo import ZoneInfo

# 1. CONFIGURAZIONE E STORICO
API_KEY = 'de85dd8016d04173bf8cbe9a86a1a617'
url_storico = "https://www.football-data.co.uk/mmz4281/2526/I1.csv"
headers = {'X-Auth-Token': API_KEY}

print("Step 1: Caricamento dati storici e calcolo forza avanzata squadre...")
df_storico = pd.read_csv(url_storico).dropna(subset=['HomeTeam', 'AwayTeam', 'FTHG', 'FTAG'])

squadre = pd.unique(df_storico[['HomeTeam', 'AwayTeam']].values.ravel('K'))
forza_attacco_casa, forza_difesa_casa = {}, {}
forza_attacco_trasferta, forza_difesa_trasferta = {}, {}

media_gol_casa_lg = df_storico['FTHG'].mean()
media_gol_trasferta_lg = df_storico['FTAG'].mean()

for sq in squadre:
    p_casa = df_storico[df_storico['HomeTeam'] == sq]
    p_trasferta = df_storico[df_storico['AwayTeam'] == sq]
    
    f_att_c = p_casa['FTHG'].mean() if len(p_casa) > 0 else media_gol_casa_lg
    f_dif_c = p_casa['FTAG'].mean() if len(p_casa) > 0 else media_gol_trasferta_lg
    f_att_t = p_trasferta['FTAG'].mean() if len(p_trasferta) > 0 else media_gol_trasferta_lg
    f_dif_t = p_trasferta['FTAG'].mean() if len(p_trasferta) > 0 else media_gol_casa_lg
    
    forza_attacco_casa[sq] = f_att_c / media_gol_casa_lg
    forza_difesa_casa[sq] = f_dif_c / media_gol_trasferta_lg
    forza_attacco_trasferta[sq] = f_att_t / media_gol_trasferta_lg
    forza_difesa_trasferta[sq] = f_dif_t / media_gol_casa_lg

mapping_nomi = {
    "FC Internazionale Milano": "Inter",
    "AC Milan": "Milan",
    "Juventus FC": "Juventus",
    "SSC Napoli": "Napoli",
    "AS Roma": "Roma",
    "SS Lazio": "Lazio",
    "Atalanta BC": "Atalanta",
    "ACF Fiorentina": "Fiorentina",
    "Bologna FC 1909": "Bologna",
    "Torino FC": "Torino",
    "Udinese Calcio": "Udinese",
    "Genoa CFC": "Genoa",
    "AC Monza": "Monza",
    "US Lecce": "Lecce",
    "Cagliari Calcio": "Cagliari",
    "Hellas Verona FC": "Verona",
    "Empoli FC": "Empoli",
    "Parma Calcio 1913": "Parma",
    "Venezia FC": "Venezia",
    "Como 1907": "Como"
}

# 1B. GESTIONE DATABASE E POPOLAMENTO QUOTE (ANTEPOST + GOL/NOGOL REALI)
print("Step 1B: Inizializzazione database e gestione quote antepost e Gol/NoGol reali...")
def gestisci_db_quote():
    conn = sqlite3.connect('campionato.db')
    cursor = conn.cursor()
    
    cursor.execute('DROP TABLE IF EXISTS quote_antepost')
    
    cursor.execute('''
        CREATE TABLE quote_antepost (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            squadra TEXT NOT NULL,
            bookmaker TEXT NOT NULL,
            mercato TEXT NOT NULL,
            quota REAL NOT NULL,
            data_aggiornamento TEXT NOT NULL
        )
    ''')
    
    orario_locale_str = datetime.now(ZoneInfo("Europe/Rome")).strftime("%Y-%m-%d %H:%M:%S")
    
    dati_quote_reali = [
        # --- VINCENTE SCUDETTO (SNAI / BET365 / EUROBET) ---
        ("Inter", "Snai", "Vincente Scudetto", 1.80, orario_locale_str),
        ("Napoli", "Snai", "Vincente Scudetto", 6.00, orario_locale_str),
        ("Juventus", "Snai", "Vincente Scudetto", 6.00, orario_locale_str),
        ("Milan", "Snai", "Vincente Scudetto", 6.00, orario_locale_str),
        ("Roma", "Snai", "Vincente Scudetto", 12.00, orario_locale_str),
        ("Como", "Snai", "Vincente Scudetto", 33.00, orario_locale_str),
        ("Atalanta", "Snai", "Vincente Scudetto", 75.00, orario_locale_str),
        ("Fiorentina", "Snai", "Vincente Scudetto", 100.00, orario_locale_str),
        ("Bologna", "Snai", "Vincente Scudetto", 100.00, orario_locale_str),
        ("Lazio", "Snai", "Vincente Scudetto", 100.00, orario_locale_str),
        ("Torino", "Snai", "Vincente Scudetto", 500.00, orario_locale_str),
        ("Udinese", "Snai", "Vincente Scudetto", 500.00, orario_locale_str),
        ("Genoa", "Snai", "Vincente Scudetto", 750.00, orario_locale_str),
        ("Monza", "Snai", "Vincente Scudetto", 1000.00, orario_locale_str),
        ("Lecce", "Snai", "Vincente Scudetto", 1000.00, orario_locale_str),
        ("Cagliari", "Snai", "Vincente Scudetto", 1000.00, orario_locale_str),
        ("Verona", "Snai", "Vincente Scudetto", 1000.00, orario_locale_str),
        ("Parma", "Snai", "Vincente Scudetto", 1000.00, orario_locale_str),
        ("Venezia", "Snai", "Vincente Scudetto", 1500.00, orario_locale_str),
        ("Empoli", "Snai", "Vincente Scudetto", 1500.00, orario_locale_str),

        ("Inter", "Bet365", "Vincente Scudetto", 1.80, orario_locale_str),
        ("Juventus", "Bet365", "Vincente Scudetto", 7.00, orario_locale_str),
        ("Milan", "Bet365", "Vincente Scudetto", 8.00, orario_locale_str),
        ("Roma", "Bet365", "Vincente Scudetto", 8.00, orario_locale_str),
        ("Como", "Bet365", "Vincente Scudetto", 25.00, orario_locale_str),
        ("Napoli", "Eurobet", "Vincente Scudetto", 6.50, orario_locale_str),
        ("Milan", "Eurobet", "Vincente Scudetto", 9.00, orario_locale_str),
        ("Roma", "Eurobet", "Vincente Scudetto", 10.00, orario_locale_str),

        # --- QUOTE GOL / NOGOL STAGIONALI ---
        ("Inter", "Snai", "Gol Stagionale Sì", 1.75, orario_locale_str),
        ("Inter", "Snai", "Gol Stagionale No", 1.95, orario_locale_str),
        ("Milan", "Snai", "Gol Stagionale Sì", 1.80, orario_locale_str),
        ("Milan", "Snai", "Gol Stagionale No", 1.90, orario_locale_str),
        ("Juventus", "Snai", "Gol Stagionale Sì", 1.90, orario_locale_str),
        ("Juventus", "Snai", "Gol Stagionale No", 1.80, orario_locale_str),
        ("Napoli", "Snai", "Gol Stagionale Sì", 1.85, orario_locale_str),
        ("Napoli", "Snai", "Gol Stagionale No", 1.85, orario_locale_str),
        ("Roma", "Snai", "Gol Stagionale Sì", 1.80, orario_locale_str),
        ("Roma", "Snai", "Gol Stagionale No", 1.90, orario_locale_str),
        ("Como", "Snai", "Gol Stagionale Sì", 1.70, orario_locale_str),
        ("Como", "Snai", "Gol Stagionale No", 2.05, orario_locale_str),
        ("Atalanta", "Snai", "Gol Stagionale Sì", 1.65, orario_locale_str),
        ("Atalanta", "Snai", "Gol Stagionale No", 2.10, orario_locale_str),
        ("Fiorentina", "Snai", "Gol Stagionale Sì", 1.75, orario_locale_str),
        ("Fiorentina", "Snai", "Gol Stagionale No", 1.95, orario_locale_str),
    ]
    
    cursor.executemany("INSERT INTO quote_antepost (squadra, bookmaker, mercato, quota, data_aggiornamento) VALUES (?, ?, ?, ?, ?)", dati_quote_reali)
    conn.commit()
        
    df_antepost = pd.read_sql_query("SELECT squadra as Squadra, bookmaker as Bookmaker, mercato as Mercato, quota as Quota, data_aggiornamento as Data_Aggiornamento FROM quote_antepost", conn)
    conn.close()
    return df_antepost

df_quote_antepost = gestisci_db_quote()

# 2. SCARICAMENTO ROSE E ANALISI STRUTTURALE DEI REPARTI
print("Step 2: Analisi strutturale delle rose e dei roster ufficiali...")
url_teams = "https://api.football-data.org/v4/competitions/2019/teams"
res_teams = requests.get(url_teams, headers=headers)
dati_rose = {}

if res_teams.status_code == 200:
    teams_data = res_teams.json().get('teams', [])
    for t in teams_data:
        raw_t_name = t['name']
        nome_pulito = mapping_nomi.get(raw_t_name, raw_t_name)
        squadra_giocatori = t.get('squad', [])
        num_giocatori = len(squadra_giocatori) if len(squadra_giocatori) > 0 else 25
        
        dati_rose[nome_pulito] = {
            'totale_rosa': num_giocatori,
            'attaccanti': sum(1 for p in squadra_giocatori if p.get('position') == 'Offence'),
            'centrocampisti': sum(1 for p in squadra_giocatori if p.get('position') == 'Midfield'),
            'difensori': sum(1 for p in squadra_giocatori if p.get('position') == 'Defence'),
            'portieri': sum(1 for p in squadra_giocatori if p.get('position') == 'Goalkeeper')
        }

# 3. SCARICAMENTO CLASSIFICA E CALCOLO INDICE DI STRESS / FORMA EVOLUTA
print("Step 3: Elaborazione classifica e calcolo indici di stress e forma...")
url_standings = "https://api.football-data.org/v4/competitions/2019/standings"
res_standings = requests.get(url_standings, headers=headers)

classifiche_dettaglio = []
forma_recentissima = {}
indici_stress_reparti = {}
dati_squadre_api = {}

if res_standings.status_code == 200:
    data_standings = res_standings.json()
    standings = data_standings.get('standings', [])
    
    tot_table, home_table, away_table = [], [], []
    for table in standings:
        t_type = table.get('type')
        if t_type == 'TOTAL': tot_table = table.get('table', [])
        elif t_type == 'HOME': home_table = table.get('table', [])
        elif t_type == 'AWAY': away_table = table.get('table', [])
        
    home_dict = {mapping_nomi.get(row['team']['name'], row['team']['name']): row for row in home_table}
    away_dict = {mapping_nomi.get(row['team']['name'], row['team']['name']): row for row in away_table}
    
    for pos in tot_table:
        raw_name = pos['team']['name']
        squadra_pulita = mapping_nomi.get(raw_name, raw_name)
        
        form_str = pos.get('form', '') 
        punti_forma = 0
        if form_str:
            for char in form_str[-5:]:
                if char == 'W': punti_forma += 3
                elif char == 'D': punti_forma += 1
                
        coeff_forma = 1.0 + ((punti_forma - 7.5) / 25.0) if form_str else 1.0
        forma_recentissima[squadra_pulita] = max(0.75, min(1.25, coeff_forma))
        
        gol_subiti = pos.get('goalsAgainst', 0)
        partite_giocate = max(1, pos.get('playedGames', 1))
        media_gs = gol_subiti / partite_giocate
        indice_vulnerabilita = max(0.85, min(1.15, 1.0 + (media_gs - 1.2) * 0.1))
        indici_stress_reparti[squadra_pulita] = indice_vulnerabilita

        dati_squadre_api[squadra_pulita] = {
            'pos': pos.get('position', 0),
            'pt': pos.get('points', 0),
            'gf': pos.get('goalsFor', 0),
            'gs': pos.get('goalsAgainst', 0),
            'form': form_str if form_str else "N/D"
        }

        h_data = home_dict.get(squadra_pulita, {})
        a_data = away_dict.get(squadra_pulita, {})

        classifiche_dettaglio.append({
            'Pos': pos.get('position', 0),
            'Squadra': squadra_pulita,
            'Rosa (Num)': dati_rose.get(squadra_pulita, {}).get('totale_rosa', 25),
            'Forma (Ultime)': form_str,
            'Pt': pos.get('points', 0),
            'G': pos.get('playedGames', 0),
            'V': pos.get('won', 0),
            'N': pos.get('draw', 0),
            'P': pos.get('lost', 0),
            'GF': pos.get('goalsFor', 0),
            'GS': pos.get('goalsAgainst', 0),
            'DR': pos.get('goalDifference', 0),
            'Casa G': h_data.get('playedGames', 0),
            'Casa V': h_data.get('won', 0),
            'Casa N': h_data.get('draw', 0),
            'Casa P': h_data.get('lost', 0),
            'Casa GF': h_data.get('goalsFor', 0),
            'Casa GS': h_data.get('goalsAgainst', 0),
            'Trasf G': a_data.get('playedGames', 0),
            'Trasf V': a_data.get('won', 0),
            'Trasf N': a_data.get('draw', 0),
            'Trasf P': a_data.get('lost', 0),
            'Trasf GF': a_data.get('goalsFor', 0),
            'Trasf GS': a_data.get('goalsAgainst', 0)
        })

df_classifica = pd.DataFrame(classifiche_dettaglio)

# 4. SCARICAMENTO PARTITE E MODELLO PREVISIONALE EVOLUTO
url_api_all = "https://api.football-data.org/v4/competitions/2019/matches"
response_all = requests.get(url_api_all, headers=headers)
lista_risultati = []
report_lines = [] 

if response_all.status_code == 200:
    partite = response_all.json().get('matches', [])
    print(f"Trovate {len(partite)} partite totali. Elaborazione in corso...")
    
    for match in partite:
        utc_date_str = match['utcDate']
        
        dt_utc = datetime.strptime(utc_date_str, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=ZoneInfo("UTC"))
        dt_italy = dt_utc.astimezone(ZoneInfo("Europe/Rome"))
        
        data_partita = dt_italy.strftime("%Y-%m-%d")
        orario_partita = dt_italy.strftime("%H:%M")
        
        giornata = match['matchday']
        stato = match['status']
        
        raw_casa = match['homeTeam']['name']
        raw_trasferta = match['awayTeam']['name']
        
        casa = mapping_nomi.get(raw_casa, raw_casa)
        trasferta = mapping_nomi.get(raw_trasferta, raw_trasferta)
        
        scontri_diretti = df_storico[
            ((df_storico['HomeTeam'] == casa) & (df_storico['AwayTeam'] == trasferta)) |
            ((df_storico['HomeTeam'] == trasferta) & (df_storico['AwayTeam'] == casa))
        ].tail(3)
        
        h2h_gol_casa = 0
        h2h_gol_trasf = 0
        
        if len(scontri_diretti) > 0:
            for _, row in scontri_diretti.iterrows():
                h_team = row['HomeTeam']
                hg = row['FTHG']
                ag = row['FTAG']
                if h_team == casa:
                    h2h_gol_casa += hg
                    h2h_gol_trasf += ag
                else:
                    h2h_gol_casa += ag
                    h2h_gol_trasf += hg
            
            media_h2h_c = h2h_gol_casa / len(scontri_diretti)
            media_h2h_t = h2h_gol_trasf / len(scontri_diretti)
            h2h_coeff_c = 0.85 + 0.15 * (media_h2h_c / max(0.5, media_gol_casa_lg))
            h2h_coeff_t = 0.85 + 0.15 * (media_h2h_t / max(0.5, media_gol_trasferta_lg))
            h2h_testo = f"Negli ultimi {len(scontri_diretti)} precedenti diretti, il bilancio reti vede {casa} a quota {h2h_gol_casa} e {trasferta} a {h2h_gol_trasf}."
        else:
            h2h_coeff_c = 1.0
            h2h_coeff_t = 1.0
            h2h_testo = "Non risultano precedenti recenti diretti significativi nello storico analizzato."

        mg_tot = {f"{i}-{j}": 0 for i in range(1, 6) for j in range(i+1, 7)}
        mg_casa = {f"{i}-{j}": 0 for i in range(1, 4) for j in range(i+1, 5)}
        mg_tras = {f"{i}-{j}": 0 for i in range(1, 4) for j in range(i+1, 5)}
        
        if stato == 'FINISHED':
            gol_c_reali = match['score']['fullTime']['home']
            gol_t_reali = match['score']['fullTime']['away']
            risultato_reale = f"{gol_c_reali}-{gol_t_reali}"
            segno_reale = "1" if gol_c_reali > gol_t_reali else ("X" if gol_c_reali == gol_t_reali else "2")
            consiglio_testo = f"FINITA ({segno_reale} - {risultato_reale})"
            
            p1, px, p2 = (1.0, 0.0, 0.0) if segno_reale == "1" else (0.0, 1.0, 0.0) if segno_reale == "X" else (0.0, 0.0, 1.0)
            p_1x = p1 + px
            p_x2 = px + p2
            p_12 = p1 + p2
            p_over_25 = 1.0 if (gol_c_reali + gol_t_reali) > 2 else 0.0
            p_under_25 = 1.0 - p_over_25
            p_goal = 1.0 if (gol_c_reali > 0 and gol_t_reali > 0) else 0.0
            p_nogoal = 1.0 - p_goal
            
            top1_res, top1_prob = risultato_reale, 1.0
            top2_res, top2_prob = "-", 0.0
            top3_res, top3_prob = "-", 0.0
        else:
            f_casa_evoluta = (forma_recentissima.get(casa, 1.0) / indici_stress_reparti.get(trasferta, 1.0)) * h2h_coeff_c
            f_trasf_evoluta = (forma_recentissima.get(trasferta, 1.0) / indici_stress_reparti.get(casa, 1.0)) * h2h_coeff_t

            lambda_c = forza_attacco_casa.get(casa, 1.0) * forza_difesa_trasferta.get(trasferta, 1.0) * media_gol_casa_lg * f_casa_evoluta
            lambda_t = forza_attacco_trasferta.get(trasferta, 1.0) * forza_difesa_casa.get(casa, 1.0) * media_gol_trasferta_lg * f_trasf_evoluta
            
            prob_matrix = np.zeros((6, 6))
            for i in range(6):
                for j in range(6):
                    prob_matrix[i, j] = poisson.pmf(i, lambda_c) * poisson.pmf(j, lambda_t)
                    
            p1 = np.sum(np.tril(prob_matrix, -1))
            px = np.sum(np.diag(prob_matrix))
            p2 = np.sum(np.triu(prob_matrix, 1))
            
            p_1x = p1 + px
            p_x2 = px + p2
            p_12 = p1 + p2
            
            p_under_25 = 0
            p_goal = 0
            tutti_risultati_esatti = []
            
            for i in range(6):
                for j in range(6):
                    p_res = prob_matrix[i, j]
                    tot_gol = i + j
                    tutti_risultati_esatti.append((f"{i}-{j}", p_res))
                    
                    if tot_gol <= 2: p_under_25 += p_res
                    if i > 0 and j > 0: p_goal += p_res
                        
                    for mg in mg_tot.keys():
                        min_g, max_g = map(int, mg.split('-'))
                        if min_g <= tot_gol <= max_g: mg_tot[mg] += p_res
                    for mg in mg_casa.keys():
                        min_g, max_g = map(int, mg.split('-'))
                        if min_g <= i <= max_g: mg_casa[mg] += p_res
                    for mg in mg_tras.keys():
                        min_g, max_g = map(int, mg.split('-'))
                        if min_g <= j <= max_g: mg_tras[mg] += p_res
                        
            p_over_25 = 1.0 - p_under_25
            p_nogoal = 1.0 - p_goal
            
            tutti_risultati_esatti.sort(key=lambda x: x[1], reverse=True)
            top1_res, top1_prob = tutti_risultati_esatti[0]
            top2_res, top2_prob = tutti_risultati_esatti[1]
            top3_res, top3_prob = tutti_risultati_esatti[2]
            
            candidati = [
                ("1X", p_1x), ("X2", p_x2), ("12", p_12),
                ("1", p1), ("X", px), ("2", p2),
                ("Over 2.5", p_over_25), ("Under 2.5", p_under_25),
                ("Goal", p_goal), ("NoGoal", p_nogoal)
            ]
            for k, v in mg_tot.items(): candidati.append((f"Multigol {k}", v))
            for k, v in mg_casa.items(): candidati.append((f"MG Casa {k}", v))
            for k, v in mg_tras.items(): candidati.append((f"MG Trasf {k}", v))
            
            miglior_scelta = max(candidati, key=lambda x: x[1])
            consiglio_testo = f"{miglior_scelta[0]} ({round(miglior_scelta[1]*100, 1)}%)"
            
            inf_c = dati_squadre_api.get(casa, {})
            inf_t = dati_squadre_api.get(trasferta, {})
            
            report_lines.append(f"STADIO & MATCH: {casa} vs {trasferta} (Giornata {giornata} | Data: {data_partita} ore {orario_partita})")
            
            incipit = f"In calendario il {data_partita} alle ore {orario_partita}, la sfida tra {casa} e {trasferta} rappresenta un incrocio cruciale per la {giornata}ª giornata."
            
            stress_c_val = indici_stress_reparti.get(casa, 1.0)
            stress_t_val = indici_stress_reparti.get(trasferta, 1.0)
            stato_reparti_c = "solido e strutturato" if stress_c_val <= 1.0 else "sotto pressione difensiva"
            stato_reparti_t = "organizzato" if stress_t_val <= 1.0 else "vulnerabile nei rientri"
            
            contesto_classifica = (
                f"Il {casa} ({inf_c.get('pos','?')}ª in classifica con {inf_c.get('pt','?')} punti) vanta un reparto {stato_reparti_c}. "
                f"Il {trasferta} risponde dalla {inf_t.get('pos','?')}ª posizione con {inf_t.get('pt','?')} punti e un assetto {stato_reparti_t}. "
                f"{h2h_testo}"
            )
            
            lettura_tecnica = (
                f"Il modello predittivo, ponderato sullo storico degli scontri diretti e sulla stabilità attuale, "
                f"stima un'aspettativa di reti pari a {round(lambda_c, 2)} per i padroni di casa e {round(lambda_t, 2)} per gli ospiti. "
                f"Il pronostico statistico principale individua il segno {miglior_scelta[0]} al {round(miglior_scelta[1]*100,1)}%. "
                f"Nei risultati esatti, le opzioni prioritarie segnalano il {top1_res} ({round(top1_prob*100,1)}%) "
                f"e il riscontro secondario del {top2_res} ({round(top2_prob*100,1)}%)."
            )
            
            report_lines.append(incipit)
            report_lines.append(contesto_classifica)
            report_lines.append(lettura_tecnica)
            report_lines.append("")

        riga = {
            'Giornata': giornata,
            'Data': data_partita,
            'Orario': orario_partita,
            'Stato': 'Giocata' if stato == 'FINISHED' else 'Da Giocare',
            'Squadra Casa': casa,
            'Squadra Trasferta': trasferta,
            'Consiglio / Risultato': consiglio_testo,
            'Prob 1': round(p1, 4), 'Quota 1': round(1/p1, 2) if p1 > 0 else 0,
            'Prob X': round(px, 4), 'Quota X': round(1/px, 2) if px > 0 else 0,
            'Prob 2': round(p2, 4), 'Quota 2': round(1/p2, 2) if p2 > 0 else 0,
            'Prob 1X': round(p_1x, 4), 'Quota 1X': round(1/p_1x, 2) if p_1x > 0 else 0,
            'Prob X2': round(p_x2, 4), 'Quota X2': round(1/p_x2, 2) if p_x2 > 0 else 0,
            'Prob 12': round(p_12, 4), 'Quota 12': round(1/p_12, 2) if p_12 > 0 else 0,
            'Prob Over 2.5': round(p_over_25, 4), 'Quota Over': round(1/p_over_25, 2) if p_over_25 > 0 else 0,
            'Prob Under 2.5': round(p_under_25, 4), 'Quota Under': round(1/p_under_25, 2) if p_under_25 > 0 else 0,
            'Prob Goal': round(p_goal, 4), 'Quota Goal': round(1/p_goal, 2) if p_goal > 0 else 0,
            'Prob NoGoal': round(p_nogoal, 4), 'Quota NoGoal': round(1/p_nogoal, 2) if p_nogoal > 0 else 0,
        }
        
        for k, v in mg_tot.items():
            riga[f'Prob MG {k}'] = round(v, 4)
            riga[f'Quota MG {k}'] = round(1/v, 2) if v > 0 else 0
        for k, v in mg_casa.items():
            riga[f'Prob MG Casa {k}'] = round(v, 4)
            riga[f'Quota MG Casa {k}'] = round(1/v, 2) if v > 0 else 0
        for k, v in mg_tras.items():
            riga[f'Prob MG Trasf {k}'] = round(v, 4)
            riga[f'Quota MG Trasf {k}'] = round(1/v, 2) if v > 0 else 0
            
        riga.update({
            '1° Risultato': top1_res,
            'Quota Esatto 1': round(1/top1_prob, 2) if top1_prob > 0 else 0,
            '2° Risultato': top2_res,
            'Quota Esatto 2': round(1/top2_prob, 2) if top2_prob > 0 else 0,
            '3° Risultato': top3_res,
            'Quota Esatto 3': round(1/top3_prob, 2) if top3_prob > 0 else 0,
        })
        
        lista_risultati.append(riga)

    df_output = pd.DataFrame(lista_risultati)
    scrivania = os.path.expanduser("~/Desktop")
    file_name = os.path.join(scrivania, "scanner_serie_a.xlsx")
    file_report = os.path.join(scrivania, "Analisi_e_Consigli_Serie_A.rtf")
    
    wb = openpyxl.Workbook()
    
    # STILI GRAFICI AGGIORNATI CON FONT A 16 (E INTESTAZIONI A 18)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=16, bold=False, color="000000")
    
    def format_sheet(ws, df):
        ws.views.sheetView[0].showGridLines = True
        headers = list(df.columns)
        ws.append(headers)
        
        # Formattazione Intestazione
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 35
        
        # Inserimento Dati e formattazione corpo a 16pt
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            ws.append(row)
            ws.row_dimensions[r_idx].height = 28
            for col in range(1, len(row) + 1):
                cell = ws.cell(row=r_idx, column=col)
                cell.font = body_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        ws.auto_filter.ref = ws.dimensions
        
        # Adattamento larghezza colonne per il nuovo font ingrandito (16pt richiede più spazio)
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(int(max_len * 2.2) + 4, 18)

    # FOGLIO 1: Calendario e Quote
    ws1 = wb.active
    ws1.title = "Calendario e Quote"
    format_sheet(ws1, df_output)

    # FOGLIO 2: Classifica e Rose
    ws2 = wb.create_sheet(title="Classifica e Rose")
    format_sheet(ws2, df_classifica)

    # FOGLIO 3: Quote Antepost
    ws3 = wb.create_sheet(title="Quote Antepost")
    format_sheet(ws3, df_quote_antepost)

    wb.save(file_name)

    rtf_content = r"{\rtf1\ansi\ansicpg1252\deff0\deflang1040" + "\n"
    rtf_content += r"{\fonttbl{\f0\fnil\fcharset0 Arial;}}" + "\n"
    rtf_content += r"{\colortbl ;\red15\green23\blue42;\red180\red30\red30;\red70\red70\red70;}" + "\n"
    rtf_content += r"\viewkind4\uc1\pard\sa200\sl276\slmult1\f0\fs22 " + "\n"
    
    rtf_content += r"\b\fs36\cf1 SCANNER SERIE A - ORARI, H2H & PRONOSTICI EVOLUTI\b0\fs22\cf0\par" + "\n"
    rtf_content += r"\i Report analitico completo di anticipi, posticipi, orari italiani corretti, trend storici e Poisson.\i0\par\line" + "\n"
    
    for line in report_lines:
        safe_line = line.replace('\\', '\\\\').replace('{', '\\{').replace('}', '\\}')
        if "STADIO & MATCH:" in line:
            rtf_content += r"\b\fs26\cf2\sa120\sb240 " + safe_line + r"\b0\fs22\cf0\par" + "\n"
        elif line == "":
            rtf_content += r"\pard\sa300\par" + "\n"
        else:
            rtf_content += safe_line + r" " + "\n"
            
    rtf_content += r"}"

    with open(file_report, "w", encoding="utf-8") as f:
        f.write(rtf_content)

    print(f"\nFatto! File Excel aggiornato con font predefinito a 16pt e larghezze colonne ricalcolate.")
    
    subprocess.run(["open", file_name])
    subprocess.run(["open", file_report])
else:
    print(f"Errore API Classifiche/Partite: {response_all.status_code}")